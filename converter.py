import pandas as pd
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import warnings

warnings.filterwarnings('ignore')

# --- 유틸리티 함수 ---

def translate_korean_date_period(text):
    """한글 날짜 기간을 숫자 형식으로 변환"""
    if pd.isna(text) or text is None:
        return text

    text_str = str(text).strip()

    try:
        pattern = r'(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일\s*(?:부터|~|-)\s*(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일\s*(?:까지|)'
        match = re.search(pattern, text_str)

        if match:
            start_year, start_month, start_day, end_year, end_month, end_day = match.groups()
            start_date = f"{start_year}.{start_month.zfill(2)}.{start_day.zfill(2)}"
            end_date = f"{end_year}.{end_month.zfill(2)}.{end_day.zfill(2)}"
            return f"{start_date} - {end_date}"

        single_pattern = r'(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일'
        single_match = re.search(single_pattern, text_str)

        if single_match:
            year, month, day = single_match.groups()
            converted_date = f"{year}.{month.zfill(2)}.{day.zfill(2)}"
            return re.sub(single_pattern, converted_date, text_str)

        return text_str

    except (ValueError, TypeError):
        return text_str

def is_total_row(row_data):
    """합계 행인지 확인"""
    if not row_data or len(row_data) < 2:
        return False

    try:
        description = str(row_data[1]).strip().lower() if pd.notna(row_data[1]) else ""
    except IndexError:
        return False

    if not description:
        return False

    total_keywords = [
        'total', 'subtotal', 'sum', 'balance', 'carry forward', '계', '합계', '소계',
        '누계', '월계', '총계', '잔액', '이월', '전기이월', '당기이월', '기말잔액',
        'beginning balance', 'monthly total', 'cumulative total', 'ending balance'
    ]

    return any(keyword in description for keyword in total_keywords)

# --- 메인 변환기 클래스 ---

class DouzoneConverter:
    def __init__(self):
        self.columns = {
            '날짜': 'Date', '적    요    란': 'Description', '코드': 'Code',
            '거래처': 'Customer/Vendor', '차   변': 'Debit', '대   변': 'Credit',
            '잔   액': 'Balance'
        }
        self.translations = {
            '전기이월': 'Beginning Balance', '월계': 'Monthly total', '누계': 'Cumulative total',
            '계   정   별   원   장': 'General Ledger', '계정별원장': 'General Ledger',
            '계정 별 원장': 'General Ledger', '계 정 별 원 장': 'General Ledger',
            '총계정원장': 'General Ledger', 
            '회사명:': 'Company Name : ', '계정과목': 'Account', '이월결손금': 'Accumulated Deficit',
            '[ 월         계 ]': 'Monthly total', '[ 누         계 ]': 'Cumulative total',
            '[월계]': 'Monthly total', '[누계]': 'Cumulative total'
        }
        self.setup_styles()
        # ✅ 여기에 추가
        self.reference_col_widths = {}

    def setup_styles(self):
        """스타일 관련 설정을 미리 정의"""
        self.title_font = Font(name='Arial', size=20, bold=True)
        self.header_font = Font(name='Arial', size=10, bold=True)
        self.data_font = Font(name='Arial', size=9)
        self.total_font = Font(name='Arial', size=9, bold=True)
        # ✅ 3행 전용 폰트 스타일
        self.row3_font = Font(name='Arial', size=9)

        self.center_align = Alignment(horizontal='center', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')
        self.right_align = Alignment(horizontal='right', vertical='center')

        self.thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        self.total_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        self.header_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        # 첫 번째 행 배경색 제거 - title_fill 삭제하고 None으로 설정
        self.title_fill = None

    def translate_text(self, text):
        if pd.isna(text) or text is None:
            return text

        try:
            text = str(text).strip()
            if not text:
                return text

            text = translate_korean_date_period(text)

            if re.search(r'계정.*원장|원장.*계정', text):
                return 'General Ledger'
            if re.search(r'\[\s*월\s*계\s*\]', text):
                return 'Monthly total'
            if re.search(r'\[\s*누\s*계\s*\]', text):
                return 'Cumulative total'

            if text in self.translations:
                return self.translations[text]

            for korean, english in self.translations.items():
                if korean in text:
                    text = text.replace(korean, english)

            return text
        except (ValueError, TypeError):
            return str(text) if text is not None else ""

    def standardize_date(self, date_str):
        if pd.isna(date_str) or date_str is None:
            return date_str

        try:
            date_text = str(date_str).strip()
            if not date_text:
                return date_str

            date_obj = pd.to_datetime(date_text, errors='coerce')
            if pd.notna(date_obj):
                return date_obj.strftime('%Y-%m-%d')
        except (ValueError, TypeError):
            pass

        return date_str

    def format_currency(self, value):
        if pd.isna(value) or value is None:
            return value

        try:
            if str(value).strip() == "":
                return value

            cleaned_value = str(value).replace(',', '').replace(' ', '')

            if not re.match(r'^-?\d*\.?\d+$', cleaned_value):
                return value

            num_value = float(cleaned_value)

            if num_value == 0:
                return 0
            return int(num_value) if abs(num_value - int(num_value)) < 0.01 else round(num_value, 2)

        except (ValueError, TypeError):
            return value

    def clean_sheet_name(self, name):
        name = re.sub(r'^\d+_', '', name)
        name = re.sub(r'\([^)]*\)', '', name)
        name = name.strip()
        return self.translate_text(name)

    def process_sheet(self, df, sheet_name):
        if df.empty or len(df) < 4:
            return None

        ws_data = []

        self._process_header_rows(df, ws_data)
        self._process_data_rows(df, ws_data)

        return ws_data

    def _process_header_rows(self, df, ws_data):
        translation_positions = {0: [0, 1, 2], 1: [1, 2, 3], 2: [0, 6]}

        for i in range(3):
            if i < len(df):
                try:
                    row_values = df.iloc[i].values
                    row = self._translate_header_row(i, row_values, translation_positions)
                    ws_data.append(row)
                except IndexError:
                    ws_data.append([None] * 7)

    def _translate_header_row(self, row_index, row_values, positions):
        row = []
        target_cols = positions.get(row_index, [])
        max_cols = len(row_values)

        if row_index == 0 and max_cols > 3 and pd.notna(row_values[3]):
            d1_content = str(row_values[3]).strip()
            if re.search(r'계정.*원장|원장.*계정', d1_content) or d1_content in self.translations:
                translated_title = self.translate_text(d1_content)
                print(f"    🔍 D1 셀 원본: '{d1_content}' → A1으로 이동: '{translated_title}'")
                row.append(translated_title)
                for j in range(1, max_cols):
                    row.append(None if j == 3 else row_values[j])
                return row

        for j in range(max_cols):
            cell = row_values[j]
            if j in target_cols and pd.notna(cell) and str(cell).strip():
                row.append(self.translate_text(cell))
            else:
                row.append(cell)
        return row

    def _process_data_rows(self, df, ws_data):
        if len(df) <= 3:
            return

        original_headers = df.iloc[3].values
        english_headers = [self.columns.get(str(h).strip(), str(h)) for h in original_headers]
        ws_data.append(english_headers)

        header_map = {header: idx for idx, header in enumerate(english_headers) if header}
        desc_idx = header_map.get('Description', -1)
        code_idx = header_map.get('Code', -1)
        date_idx = header_map.get('Date', -1)
        money_indices = [header_map.get(col, -1) for col in ['Debit', 'Credit', 'Balance'] if col in header_map]

        for i in range(4, len(df)):
            try:
                row_values = df.iloc[i].values
                row_data = self._process_single_data_row(row_values, desc_idx, code_idx, date_idx, money_indices)
                if any(pd.notna(x) and str(x).strip() for x in row_data if x is not None):
                    ws_data.append(row_data)
            except IndexError:
                continue

    def _process_single_data_row(self, row_values, desc_idx, code_idx, date_idx, money_indices):
        row_data = []
        for j, cell in enumerate(row_values):
            if j == desc_idx:
                cell = self.translate_text(cell)
            elif j == code_idx and pd.notna(cell):
                cell = str(cell).strip()
            elif j == date_idx:
                cell = self.standardize_date(cell)
            elif j in money_indices:
                cell = self.format_currency(cell)
            row_data.append(cell)
        return row_data

    def apply_formatting(self, ws):
        self._apply_general_formatting(ws)
        # 2행 A2:G2 병합 + D2 값 보존
        try:
            merge_range = 'A2:G2'
            value = ws['D2'].value  # D2 값 보존
            ws.merge_cells(merge_range)
            cell = ws['A2']
            cell.value = value
            cell.font = self.header_font
            cell.alignment = self.center_align
            ws.row_dimensions[2].height = 22
        except Exception as e:
            print(f"⚠️ 2행 병합 중 오류: {e}")
        # C열 (Code 열, 3번째 열) 오류 알림 무시 + 텍스트 서식 적용
        for row in ws.iter_rows(min_row=5, min_col=3, max_col=3, max_row=ws.max_row):
            for cell in row:
                if cell.value is not None:
                    cell.number_format = '@'  # 텍스트 형식

        # ✅ 3행 폰트: Arial 9pt 적용
        for cell in ws[3]:
            cell.font = self.row3_font
        print("    🎨 3행 폰트: Arial, 크기 9 적용 완료")   
    

        self._apply_total_row_formatting(ws)
        self._adjust_column_widths(ws)
        self._set_active_cell(ws)

    def _apply_general_formatting(self, ws):
        if ws.max_row > 0:
            ws.merge_cells('A1:G1')
            ws['A1'].font = self.title_font
            ws['A1'].alignment = self.center_align
            # 첫 번째 행 배경색 제거 - fill 설정 삭제
            # ws['A1'].fill = self.title_fill  # 이 줄을 주석 처리
            ws.row_dimensions[1].height = 25

        if ws.max_row >= 3:
            try:
                ws['G3'].alignment = self.right_align
            except (KeyError, IndexError):
                pass

        if ws.max_row >= 4:
            for cell in ws[4]:
                cell.font = self.header_font
                cell.alignment = self.center_align
                cell.border = self.thin_border
                cell.fill = self.header_fill

        for row_num in range(5, ws.max_row + 1):
            max_col = min(7, ws.max_column)
            for col_num in range(1, max_col + 1):
                try:
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.font = self.data_font
                    cell.border = self.thin_border
                    if col_num in [1, 3]:
                        cell.alignment = self.center_align
                    elif col_num in [5, 6, 7]:
                        cell.alignment = self.right_align
                        cell.number_format = '#,##0'
                    else:
                        cell.alignment = self.left_align
                except (KeyError, IndexError):
                    continue

        # ✅ 틀고정: 5행을 기준으로 위쪽 고정
        ws.freeze_panes = 'A5'
        print("    📌 틀고정: 5행 위쪽 고정 완료")        

    def _apply_total_row_formatting(self, ws):
        total_blocks = self._find_total_blocks(ws)

        for start_row, end_row in total_blocks:
            for row_num in range(start_row, end_row + 1):
                max_col = min(7, ws.max_column)
                for col_num in range(1, max_col + 1):
                    try:
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.font = self.total_font
                        cell.fill = self.total_fill
                        border = Border(
                            left=Side(style='thin') if col_num == 1 else None,
                            right=Side(style='thin') if col_num == max_col else None,
                            top=Side(style='thin') if row_num == start_row else None,
                            bottom=Side(style='thin') if row_num == end_row else None
                        )
                        cell.border = border
                    except (KeyError, IndexError):
                        continue

    def _find_total_blocks(self, ws):
        total_blocks = []
        current_block_start = None
        for row_num in range(5, ws.max_row + 1):
            try:
                row_data = [ws.cell(row=row_num, column=col).value for col in range(1, min(8, ws.max_column + 1))]
                if is_total_row(row_data):
                    if current_block_start is None:
                        current_block_start = row_num
                elif current_block_start is not None:
                    total_blocks.append((current_block_start, row_num - 1))
                    current_block_start = None
            except (KeyError, IndexError):
                continue
        if current_block_start is not None:
            total_blocks.append((current_block_start, ws.max_row))
        return total_blocks

    def _adjust_column_widths(self, ws):
        # ✅ 1. 기준 너비가 있다면 그대로 적용
        if self.reference_col_widths:
            for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                ws.column_dimensions[col_letter].width = self.reference_col_widths[col_letter]
            print("    📏 열 넓이: 'bank deposits' 기준으로 적용 완료")
            return

        # ✅ 2. 기준 없을 경우 고정 너비로 설정
        print("    📏 열 너비 고정값 적용 중...")

        fixed_widths = {
            'A': 6,
            'B': 45,
            'C': 8,
            'D': 25,
            'E': 13,
            'F': 13,
            'G': 13
        }

        for col_letter, width in fixed_widths.items():
            ws.column_dimensions[col_letter].width = width
            print(f"    📏 {col_letter}열 너비 고정: {width}")

        print("    ✅ 열 너비 고정 적용 완료")

    def _set_active_cell(self, ws):
        try:
            last_data_row = ws.max_row
            target_row = last_data_row + 4
            active_cell = f"G{target_row}"
            ws.sheet_view.selection[0].activeCell = active_cell
            ws.sheet_view.selection[0].sqref = active_cell
            print(f"    📍 액티브 셀 설정: {active_cell} (마지막 데이터: {last_data_row}행)")
        except (AttributeError, IndexError):
            print("    ⚠️ 액티브 셀 설정 실패")
            try:
                ws.sheet_view.selection[0].activeCell = "G10"
                ws.sheet_view.selection[0].sqref = "G10"
            except (AttributeError, IndexError):
                pass

    def convert(self, input_file, output_file, english_company_name=None):
        try:
            with open(output_file, 'a'):
                os.utime(output_file, None)
        except PermissionError:
            print(f"❌ 파일이 사용 중입니다: {output_file}")
            return False
        except IOError as e:
            print(f"❌ 출력 파일 접근 오류: {e}")
            return False

        print(f"🔄 변환 중: {input_file}")

        try:
            excel_file = pd.ExcelFile(input_file)
        except FileNotFoundError:
            print(f"❌ 파일을 찾을 수 없습니다: {input_file}")
            return False
        except Exception as e:
            print(f"❌ Excel 파일 읽기 실패: {e}")
            return False

        try:
            wb = Workbook()
            wb.remove(wb.active)
        except Exception as e:
            print(f"❌ 워크북 생성 실패: {e}")
            return False

        processed_sheets = 0
        for sheet_name in excel_file.sheet_names:
            try:
                print(f"  📋 처리 중: {sheet_name}")
                df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
                ws_data = self.process_sheet(df, sheet_name)

                if ws_data:
                    ws = wb.create_sheet(title=sheet_name)
                    for row_data in ws_data:
                        ws.append(row_data)

                    # ✅ 선택된 영문 회사명을 A3 셀에 삽입
                    if english_company_name:
                        ws['A3'] = f"Company Name : {english_company_name}"    

                    self.apply_formatting(ws)
                    processed_sheets += 1

                    # ✅ 'bank deposits' 시트면 열 너비 저장
                    if 'bank deposits' in sheet_name.lower():
                        self.reference_col_widths = {
                            col_letter: ws.column_dimensions[col_letter].width
                            for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G']
                        }
                else:
                    print(f"    ⚠️ 빈 시트 건너뜀: {sheet_name}")
            except Exception as e:
                print(f"    ❌ 시트 처리 실패 ({sheet_name}): {e}")
                continue

        if processed_sheets == 0:
            print("❌ 처리할 수 있는 시트가 없습니다.")
            return False

        try:
            wb.save(output_file)
            print(f"✅ 변환 완료: {output_file} ({processed_sheets}/{len(excel_file.sheet_names)} 시트 처리됨)")
            return True
        except PermissionError:
            print(f"❌ 파일 저장 권한 오류: {output_file}")
            return False
        except Exception as e:
            print(f"❌ 파일 저장 실패: {e}")
            return False